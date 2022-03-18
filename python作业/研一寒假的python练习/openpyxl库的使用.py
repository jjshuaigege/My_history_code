import openpyxl as pxl
book=pxl.load_workbook("../../上课信息（下）.xlsx")
sheet_0=book.worksheets[0]
print(sheet_0.title)
print(sheet_0.min_row,sheet_0.max_row)
print(sheet_0.min_column,sheet_0.max_column)